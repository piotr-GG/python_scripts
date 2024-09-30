import re
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, numbers


@dataclass
class CAN_dbc:
    name: str
    dbc_path: str


DBC_PATHS = [
    CAN_dbc(name="CANFD2",
            dbc_path=r"sample_dbc.dbc")
]

PARSED_OBJECTS = {
    # "CAN_MSG": [],
    # "COMMENTS": [],
    # "VALUES": []
}

CAN_MSG_REGEX = r"^(BO_\s.*\n)((?:\s+SG_.*)*)"
BU_REGEX = r"^(BU_:\s(.*))"
BO_TX_BU_REGEX = r"^(?:BO_TX_BU_\s)(.*);$"
CM_REGEX = r"^(?:CM_\s)(.*);$"
VAL_REGEX = r"^(?:VAL_\s)(.*);$"


@dataclass
class CAN_Message:
    msg_name: str
    msg_id: int
    msg_len: int
    sender: str
    signals: list

    def __str__(self):
        return f"CAN Message: {self.msg_name} ID: {self.msg_id} MSG length: {self.msg_len}"

    def to_row(self):
        return [self.msg_name, self.msg_id, self.msg_len, self.sender]

    @classmethod
    def get_headers(cls):
        return ["Message name", "Message ID", "Message length", "Sender"]


@dataclass
class CAN_Signal:
    signal_name: str
    bit_start: int
    bit_len: int
    endianness: str
    unsigned: bool
    scale: float
    offset: float
    min: float
    max: float
    unit: str
    receivers: list
    enums: dict = None

    def __str__(self):
        return (f"{self.signal_name} start_bit: {self.bit_start} bit length: {self.bit_len} "
                f"endianness: {self.endianness}, unsigned: {self.unsigned}, scale: {self.scale}, offset: {self.offset}, "
                f"min: {self.min}, max: {self.max}, unit: {self.unit}, receivers: {' '.join(receivers)}")

    def to_row(self):
        return [self.signal_name, self.bit_start, self.bit_len, self.endianness, self.unsigned, self.scale, self.offset,
                self.min, self.max, self.unit, " ".join(receivers)]

    @classmethod
    def get_headers(cls):
        return ["Signal name", "Start bit", "Bit length", "Endianness", "Unsigned", "Scale", "Offset", "Min", "Max",
                "Unit", "Receivers"]


def auto_adjust_column_width(sheet):
    print("Sheet name:", sheet.title)
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)) + 0.71)
            except:
                pass
        adjusted_width = max_length
        sheet.column_dimensions[column_letter].width = adjusted_width
        print("Column:", column_letter)
        print("Adjusted width:", max_length)


def write_data_to_xlsx(xlsx_path, data):
    workbook = Workbook()
    worksheet = workbook.active

    workbook.remove(worksheet)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    align_center = Alignment(horizontal="center", vertical="center")

    for can_dbc in data:
        worksheet = workbook.create_sheet(title=f"{can_dbc} messages")
        headers = CAN_Message.get_headers()
        for col_index, header_data in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_index, value=header_data)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = align_center

        raw_data = []
        for message in data[can_dbc]:
            raw_data.append(message.to_row())
        for row_index, row_data in enumerate(raw_data, start=2):
            for col_index, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=value)
                cell.border = border
        auto_adjust_column_width(worksheet)

        worksheet = workbook.create_sheet(title=f"{can_dbc} signals")
        headers = ["Message name"] + CAN_Signal.get_headers()
        for col_index, header_data in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_index, value=header_data)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = align_center

        raw_data = []
        for message in data[can_dbc]:
            for signal in message.signals:
                raw_data.append([message.msg_name] + signal.to_row())
            raw_data.append(["-"] * (len(signal.to_row()) + 1))
        for row_index, row_data in enumerate(raw_data, start=2):
            for col_index, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=value)
                cell.border = border
        auto_adjust_column_width(worksheet)

        worksheet = workbook.create_sheet(title=f"{can_dbc} enums")
        headers = ["Message name", "Enum", "Value"]
        for col_index, header_data in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_index, value=header_data)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = align_center

        raw_data = []
        for message in data[can_dbc]:
            for signal in message.signals:
                if signal.enums:
                    raw_data.append([f"{message.msg_name}::{signal.signal_name}"])
                    for val, enum in signal.enums:
                        print(val, enum)
                        raw_data.append(["", int(val.replace('"', '').strip()), enum])
        for row_index, row_data in enumerate(raw_data, start=2):
            for col_index, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=value)
                cell.border = border
        auto_adjust_column_width(worksheet)
    workbook.save(xlsx_path)


for can_dbc in DBC_PATHS:
    try:
        PARSED_OBJECTS[can_dbc.name]
    except KeyError:
        PARSED_OBJECTS[can_dbc.name] = []

    with open(can_dbc.dbc_path) as f:
        dbc_text = f.read()

    can_msgs = re.findall(CAN_MSG_REGEX, dbc_text, flags=re.MULTILINE)
    bu = re.findall(BU_REGEX, dbc_text, flags=re.MULTILINE)
    bo_tx_bu = re.findall(BO_TX_BU_REGEX, dbc_text, flags=re.MULTILINE)
    comments = re.findall(CM_REGEX, dbc_text, flags=re.MULTILINE)
    vals = re.findall(VAL_REGEX, dbc_text, flags=re.MULTILINE)

    for can_msg_data in can_msgs:
        can_msg = can_msg_data[0]
        signals = [el.replace("SG_", "") for el in can_msg_data[1].splitlines()]

        id_msg, msg_name, msg_len, sender_name = re.findall(r"^BO_\s(\d+)\s(\w+):\s(\d+)\s(\w+)", can_msg)[0]

        can_message = CAN_Message(msg_name, int(id_msg), int(msg_len), sender_name, [])

        for signal in signals:
            signal_name, signal_data = [el.strip() for el in signal.split(":")]
            signal_fields = signal_data.split(" ")

            bit_start_len_endian = signal_fields[0]
            scale_offset = signal_fields[1]
            min_max = signal_fields[2]
            unit = signal_fields[3]
            receiver = signal_fields[4]

            bit_start = bit_start_len_endian.split("|")[0]
            bit_len, endian, unsigned = re.findall(r"(\d+)@(\d+)([+-])", bit_start_len_endian.split("|")[1])[0]

            if endian:
                endianness = "little-endian"
            else:
                endianness = "big-endian"

            if unsigned[0] == "+":
                is_unsigned = True
            elif unsigned[0] == "-":
                is_unsigned = False
            else:
                raise ValueError("Unsigned[0] does not have proper value!")

            scale, offset = [el.strip().replace("(", "").replace(")", "") for el in scale_offset.split(",")]

            min_val, max_val = [el.strip().replace("[", "").replace("]", "") for el in min_max.split("|")]

            unit = unit.replace("\"", "")

            receivers = receiver.split(",")
            signal = CAN_Signal(signal_name=signal_name,
                                bit_start=int(bit_start),
                                bit_len=int(bit_len),
                                endianness=endianness,
                                unsigned=is_unsigned,
                                scale=float(scale), offset=float(offset), min=float(min_val), max=float(max_val),
                                unit=unit,
                                receivers=receivers)
            can_message.signals.append(signal)
        PARSED_OBJECTS[can_dbc.name].append(can_message)

    signal_val_dict = dict()
    for value in vals:
        pattern = r'(\d+) (\w+) (.+)'
        match = re.match(pattern, value)

        if not match:
            raise ValueError("The input string does not match the expected format.")

        msg_id = int(match.group(1))
        signal_name = match.group(2)

        enum_part = match.group(3)
        pair_pattern = r'(\d+) "([^"]+)"'
        enum_data = re.findall(pair_pattern, enum_part)

        signal_dict = signal_val_dict.get(msg_id, dict())
        signal_dict[signal_name] = enum_data
        signal_dict[msg_id] = signal_dict[signal_name]
        signal_val_dict[msg_id] = signal_dict

    for can_msg in PARSED_OBJECTS[can_dbc.name]:
        msg_signal_values = signal_val_dict.get(can_msg.msg_id)
        if msg_signal_values:
            for signal in can_msg.signals:
                signal.enums = msg_signal_values.get(signal.signal_name)

    for can_msg in PARSED_OBJECTS[can_dbc.name]:
        for signal in can_msg.signals:
            print(signal.signal_name)
            print(signal.enums)

for can_dbc in PARSED_OBJECTS:
    print("*" * 50)
    print(f"CAN DBC {can_dbc}")
    print("*" * 50)
    sig_separator = "-" * 5 + ">"
    for can_msg in PARSED_OBJECTS[can_dbc]:
        print(can_msg)
        for signal in can_msg.signals:
            print(sig_separator, signal)

write_data_to_xlsx(r"dbc_output_new.xlsx", PARSED_OBJECTS)