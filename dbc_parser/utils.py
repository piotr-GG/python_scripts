from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

from excel_handling import auto_adjust_column_width


@dataclass
class CAN_dbc:
    name: str
    dbc_path: str


@dataclass
class CAN_Message:
    msg_name: str
    msg_id: int
    msg_len: int
    sender: str
    signals: list

    def __str__(self):
        return f"CAN Message: {self.msg_name} ID: {self.msg_id} MSG length: {self.msg_len}"

    def to_row(self):
        return [self.msg_name, self.msg_id, self.msg_len, self.sender]

    @classmethod
    def get_headers(cls):
        return ["Message name", "Message ID", "Message length", "Sender"]


@dataclass
class CAN_Signal:
    signal_name: str
    bit_start: int
    bit_len: int
    endianness: str
    unsigned: bool
    scale: float
    offset: float
    min: float
    max: float
    unit: str
    receivers: list
    enums: dict = None

    def __str__(self):
        return (f"{self.signal_name} start_bit: {self.bit_start} bit length: {self.bit_len} "
                f"endianness: {self.endianness}, unsigned: {self.unsigned}, scale: {self.scale}, offset: {self.offset}, "
                f"min: {self.min}, max: {self.max}, unit: {self.unit}, receivers: {' '.join(self.receivers)}")

    def to_row(self):
        return [self.signal_name, self.bit_start, self.bit_len, self.endianness, self.unsigned, self.scale, self.offset,
                self.min, self.max, self.unit, " ".join(self.receivers)]

    @classmethod
    def get_headers(cls):
        return ["Signal name", "Start bit", "Bit length", "Endianness", "Unsigned", "Scale", "Offset", "Min", "Max",
                "Unit", "Receivers"]


def write_data_to_xlsx(xlsx_path, data):
    workbook = Workbook()
    worksheet = workbook.active

    workbook.remove(worksheet)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    align_center = Alignment(horizontal="center", vertical="center")

    for can_dbc in data:
        worksheet = workbook.create_sheet(title=f"{can_dbc} messages")
        headers = CAN_Message.get_headers()
        for col_index, header_data in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_index, value=header_data)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = align_center

        raw_data = []
        for message in data[can_dbc]:
            raw_data.append(message.to_row())
        for row_index, row_data in enumerate(raw_data, start=2):
            for col_index, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=value)
                cell.border = border
        auto_adjust_column_width(worksheet)

        worksheet = workbook.create_sheet(title=f"{can_dbc} signals")
        headers = ["Message name"] + CAN_Signal.get_headers()
        for col_index, header_data in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_index, value=header_data)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = align_center

        raw_data = []
        for message in data[can_dbc]:
            for signal in message.signals:
                raw_data.append([message.msg_name] + signal.to_row())
            raw_data.append(["-"] * (len(signal.to_row()) + 1))
        for row_index, row_data in enumerate(raw_data, start=2):
            for col_index, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=value)
                cell.border = border
        auto_adjust_column_width(worksheet)

        worksheet = workbook.create_sheet(title=f"{can_dbc} enums")
        headers = ["Message name", "Enum", "Value"]
        for col_index, header_data in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_index, value=header_data)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = align_center

        raw_data = []
        for message in data[can_dbc]:
            for signal in message.signals:
                if signal.enums:
                    raw_data.append([f"{message.msg_name}::{signal.signal_name}"])
                    for val, enum in signal.enums:
                        print(val, enum)
                        raw_data.append(["", int(val.replace('"', '').strip()), enum])
        for row_index, row_data in enumerate(raw_data, start=2):
            for col_index, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=value)
                cell.border = border
        auto_adjust_column_width(worksheet)
    workbook.save(xlsx_path)
